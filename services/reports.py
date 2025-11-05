"""
services/reports.py - Service de génération et export de rapports
Gère la création de rapports de ventes, exports PDF/Excel et rapports planifiés.
"""

from models import db, Trip, Client, User, SalesReport, Agency
from services.analytics import AnalyticsService
from datetime import datetime, timedelta, date
from typing import Dict, List, Optional, Tuple
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT


class ReportsService:
    """Service de génération de rapports professionnels"""
    
    @staticmethod
    def generate_period_report(
        agency_id: int,
        period_type: str,
        custom_start: Optional[date] = None,
        custom_end: Optional[date] = None,
        user_id: Optional[int] = None
    ) -> SalesReport:
        """
        Génère un rapport pour une période donnée
        
        Args:
            agency_id: ID de l'agence
            period_type: Type de période (daily, weekly, monthly, yearly, custom)
            custom_start: Date de début personnalisée (pour custom)
            custom_end: Date de fin personnalisée (pour custom)
            user_id: ID du vendeur (None = rapport global)
            
        Returns:
            Objet SalesReport créé
        """
        # Déterminer les dates selon le type
        end_date = date.today()
        
        if period_type == 'daily':
            start_date = end_date
        elif period_type == 'weekly':
            start_date = end_date - timedelta(days=7)
        elif period_type == 'monthly':
            start_date = end_date - timedelta(days=30)
        elif period_type == 'yearly':
            start_date = end_date - timedelta(days=365)
        elif period_type == 'custom':
            if not custom_start or not custom_end:
                raise ValueError("custom_start and custom_end required for custom reports")
            start_date = custom_start
            end_date = custom_end
        else:
            raise ValueError(f"Invalid period_type: {period_type}")
        
        # Créer le rapport via AnalyticsService
        return AnalyticsService.create_sales_report(
            agency_id=agency_id,
            report_type=period_type,
            period_start=start_date,
            period_end=end_date,
            user_id=user_id
        )
    
    @staticmethod
    def export_report_to_excel(report: SalesReport) -> io.BytesIO:
        """
        Exporte un rapport au format Excel
        
        Args:
            report: Objet SalesReport à exporter
            
        Returns:
            BytesIO contenant le fichier Excel
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Rapport de Ventes"
        
        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=16)
        
        # Titre
        ws['A1'] = f"Rapport de Ventes - {report.report_type.upper()}"
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        
        # Informations du rapport
        ws['A3'] = "Agence:"
        ws['B3'] = report.agency.name
        ws['A4'] = "Période:"
        ws['B4'] = f"{report.period_start.strftime('%d/%m/%Y')} - {report.period_end.strftime('%d/%m/%Y')}"
        ws['A5'] = "Généré le:"
        ws['B5'] = report.generated_at.strftime('%d/%m/%Y à %H:%M')
        
        if report.user_id:
            ws['A6'] = "Vendeur:"
            ws['B6'] = report.user.pseudo
        
        # Métriques principales
        row = 8
        ws[f'A{row}'] = "MÉTRIQUES PRINCIPALES"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        metrics = [
            ("Nombre de ventes", report.total_sales),
            ("Chiffre d'affaires", f"{report.total_revenue} €"),
            ("Panier moyen", f"{report.average_sale} €"),
            ("Voyages créés", report.trip_count)
        ]
        
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Top destinations
        if report.detailed_data and 'destinations' in report.detailed_data:
            row += 1
            ws[f'A{row}'] = "TOP DESTINATIONS"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws.merge_cells(f'A{row}:C{row}')
            
            row += 1
            ws[f'A{row}'] = "Rang"
            ws[f'B{row}'] = "Destination"
            ws[f'C{row}'] = "Ventes"
            for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}']]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            for i, dest in enumerate(report.detailed_data['destinations'][:10], 1):
                row += 1
                ws[f'A{row}'] = i
                ws[f'B{row}'] = dest['destination']
                ws[f'C{row}'] = dest['count']
        
        # Performance par vendeur (si rapport global)
        if not report.user_id and report.detailed_data and report.detailed_data.get('seller_breakdown'):
            row += 2
            ws[f'A{row}'] = "PERFORMANCE PAR VENDEUR"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws.merge_cells(f'A{row}:D{row}')
            
            row += 1
            ws[f'A{row}'] = "Rang"
            ws[f'B{row}'] = "Vendeur"
            ws[f'C{row}'] = "Ventes"
            ws[f'D{row}'] = "CA"
            for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}'], ws[f'D{row}']]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            sellers = sorted(
                report.detailed_data['seller_breakdown'].items(),
                key=lambda x: x[1]['revenue'],
                reverse=True
            )
            
            for i, (seller_id, data) in enumerate(sellers, 1):
                row += 1
                ws[f'A{row}'] = i
                ws[f'B{row}'] = data['pseudo']
                ws[f'C{row}'] = data['count']
                ws[f'D{row}'] = f"{data['revenue']} €"
        
        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        
        # Sauvegarder dans BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def export_report_to_pdf(report: SalesReport) -> io.BytesIO:
        """
        Exporte un rapport au format PDF
        
        Args:
            report: Objet SalesReport à exporter
            
        Returns:
            BytesIO contenant le fichier PDF
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Style personnalisé pour le titre
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Titre
        title = Paragraph(f"Rapport de Ventes - {report.report_type.upper()}", title_style)
        story.append(title)
        story.append(Spacer(1, 0.3*inch))
        
        # Informations du rapport
        info_data = [
            ['Agence:', report.agency.name],
            ['Période:', f"{report.period_start.strftime('%d/%m/%Y')} - {report.period_end.strftime('%d/%m/%Y')}"],
            ['Généré le:', report.generated_at.strftime('%d/%m/%Y à %H:%M')]
        ]
        
        if report.user_id:
            info_data.append(['Vendeur:', report.user.pseudo])
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#366092')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Métriques principales
        metrics_title = Paragraph("<b>MÉTRIQUES PRINCIPALES</b>", styles['Heading2'])
        story.append(metrics_title)
        story.append(Spacer(1, 0.1*inch))
        
        metrics_data = [
            ['Nombre de ventes', str(report.total_sales)],
            ["Chiffre d'affaires", f"{report.total_revenue} €"],
            ['Panier moyen', f"{report.average_sale} €"],
            ['Voyages créés', str(report.trip_count)]
        ]
        
        metrics_table = Table(metrics_data, colWidths=[3*inch, 3*inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F2F2F2')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 12),
            ('RIGHTPADDING', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(metrics_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Top destinations
        if report.detailed_data and 'destinations' in report.detailed_data:
            dest_title = Paragraph("<b>TOP 10 DESTINATIONS</b>", styles['Heading2'])
            story.append(dest_title)
            story.append(Spacer(1, 0.1*inch))
            
            dest_data = [['Rang', 'Destination', 'Ventes']]
            for i, dest in enumerate(report.detailed_data['destinations'][:10], 1):
                dest_data.append([str(i), dest['destination'], str(dest['count'])])
            
            dest_table = Table(dest_data, colWidths=[0.8*inch, 3.5*inch, 1.2*inch])
            dest_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(dest_table)
        
        # Performance par vendeur (si rapport global)
        if not report.user_id and report.detailed_data and report.detailed_data.get('seller_breakdown'):
            story.append(Spacer(1, 0.3*inch))
            sellers_title = Paragraph("<b>PERFORMANCE PAR VENDEUR</b>", styles['Heading2'])
            story.append(sellers_title)
            story.append(Spacer(1, 0.1*inch))
            
            sellers_data = [['Rang', 'Vendeur', 'Ventes', 'CA']]
            sellers = sorted(
                report.detailed_data['seller_breakdown'].items(),
                key=lambda x: x[1]['revenue'],
                reverse=True
            )
            
            for i, (seller_id, data) in enumerate(sellers, 1):
                sellers_data.append([
                    str(i),
                    data['pseudo'],
                    str(data['count']),
                    f"{data['revenue']} €"
                ])
            
            sellers_table = Table(sellers_data, colWidths=[0.8*inch, 2*inch, 1.2*inch, 1.5*inch])
            sellers_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(sellers_table)
        
        # Construire le PDF
        doc.build(story)
        buffer.seek(0)
        
        return buffer
    
    @staticmethod
    def export_clients_to_csv(agency_id: int, client_type: Optional[str] = None) -> io.StringIO:
        """
        Exporte la liste des clients au format CSV
        
        Args:
            agency_id: ID de l'agence
            client_type: Type de client à filtrer (nouveau, regulier, vip) ou None pour tous
            
        Returns:
            StringIO contenant le CSV
        """
        query = Client.query.filter_by(agency_id=agency_id)
        
        if client_type:
            query = query.filter_by(client_type=client_type)
        
        clients = query.all()
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # En-têtes
        writer.writerow([
            'ID',
            'Prénom',
            'Nom',
            'Email',
            'Téléphone',
            'Adresse',
            'Type',
            'Achats',
            'CA Total',
            'Dernier Achat',
            'Source',
            'Date d\'inscription'
        ])
        
        # Données
        for client in clients:
            writer.writerow([
                client.id,
                client.first_name,
                client.last_name,
                client.email or '',
                client.phone or '',
                client.address or '',
                client.client_type or 'nouveau',
                client.total_purchases or 0,
                client.total_revenue or 0,
                client.last_purchase_date.strftime('%d/%m/%Y') if client.last_purchase_date else '',
                client.source or '',
                client.created_at.strftime('%d/%m/%Y')
            ])
        
        output.seek(0)
        return output
    
    @staticmethod
    def export_clients_to_excel(agency_id: int, client_type: Optional[str] = None) -> io.BytesIO:
        """
        Exporte la liste des clients au format Excel
        
        Args:
            agency_id: ID de l'agence
            client_type: Type de client à filtrer ou None pour tous
            
        Returns:
            BytesIO contenant le fichier Excel
        """
        query = Client.query.filter_by(agency_id=agency_id)
        
        if client_type:
            query = query.filter_by(client_type=client_type)
        
        clients = query.all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Clients"
        
        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        # En-têtes
        headers = [
            'ID', 'Prénom', 'Nom', 'Email', 'Téléphone', 'Adresse',
            'Type', 'Achats', 'CA Total', 'Dernier Achat', 'Source', "Date d'inscription"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Données
        for row, client in enumerate(clients, 2):
            ws.cell(row=row, column=1, value=client.id)
            ws.cell(row=row, column=2, value=client.first_name)
            ws.cell(row=row, column=3, value=client.last_name)
            ws.cell(row=row, column=4, value=client.email or '')
            ws.cell(row=row, column=5, value=client.phone or '')
            ws.cell(row=row, column=6, value=client.address or '')
            ws.cell(row=row, column=7, value=client.client_type or 'nouveau')
            ws.cell(row=row, column=8, value=client.total_purchases or 0)
            ws.cell(row=row, column=9, value=f"{client.total_revenue or 0} €")
            ws.cell(row=row, column=10, value=client.last_purchase_date.strftime('%d/%m/%Y') if client.last_purchase_date else '')
            ws.cell(row=row, column=11, value=client.source or '')
            ws.cell(row=row, column=12, value=client.created_at.strftime('%d/%m/%Y'))
        
        # Ajuster les largeurs
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15
        
        # Sauvegarder
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def get_comparison_report(
        agency_id: int,
        period1_start: date,
        period1_end: date,
        period2_start: date,
        period2_end: date
    ) -> Dict:
        """
        Compare deux périodes
        
        Args:
            agency_id: ID de l'agence
            period1_start: Début période 1
            period1_end: Fin période 1
            period2_start: Début période 2
            period2_end: Fin période 2
            
        Returns:
            Dict avec comparaison
        """
        # Créer les deux rapports
        report1 = AnalyticsService.create_sales_report(
            agency_id, 'custom', period1_start, period1_end
        )
        report2 = AnalyticsService.create_sales_report(
            agency_id, 'custom', period2_start, period2_end
        )
        
        # Calculer les variations
        def calc_variation(old_val, new_val):
            if old_val == 0:
                return 100 if new_val > 0 else 0
            return round(((new_val - old_val) / old_val) * 100, 2)
        
        return {
            'period1': {
                'start': period1_start.strftime('%d/%m/%Y'),
                'end': period1_end.strftime('%d/%m/%Y'),
                'total_sales': report1.total_sales,
                'total_revenue': report1.total_revenue,
                'average_sale': report1.average_sale,
                'trip_count': report1.trip_count
            },
            'period2': {
                'start': period2_start.strftime('%d/%m/%Y'),
                'end': period2_end.strftime('%d/%m/%Y'),
                'total_sales': report2.total_sales,
                'total_revenue': report2.total_revenue,
                'average_sale': report2.average_sale,
                'trip_count': report2.trip_count
            },
            'variations': {
                'sales': calc_variation(report1.total_sales, report2.total_sales),
                'revenue': calc_variation(report1.total_revenue, report2.total_revenue),
                'average_sale': calc_variation(report1.average_sale, report2.average_sale),
                'trips': calc_variation(report1.trip_count, report2.trip_count)
            }
        }
